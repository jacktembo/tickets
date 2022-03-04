from openpyxl import load_workbook
file_name = input('Please enter file name. Case sensitive')
wb = load_workbook(file_name)
sheet = wb.active

bus = input('Enter a Bus Short Nane')
startting_place = sheet.cell(row=1, column=1)
destination = sheet.cell(row=1, column=1)
time = sheet.cell(row=1, column=1)
price = sheet.cell(row=1, column=1)
