from django.http import HttpResponse
from django.shortcuts import render
from django.views import View
import openpyxl
from buses.models import Route, Bus


def index(request):
    return render(request, 'index.html')

class BulkRoutesUpload(View):
    template_name = 'bulk_routes_upload'

    def get(self, request):
        buses = Bus.objects.all()
        context = {
            'buses': buses,
        }
        return render(request, 'bulk_routes_upload.html', context)

    def post(self, request):
        def get_maximum_rows(*, sheet_object):
            rows = 0
            for max_row, row in enumerate(sheet_object, 1):
                if not all(col.value is None for col in row):
                    rows += 1
            return rows

        file = request.FILES['file']
        wb = openpyxl.load_workbook(file)
        sheets = wb.sheetnames
        for bus in sheets:
            sheet = wb[bus]
            maximum_rows = get_maximum_rows(sheet_object=sheet)
            routes = []
            obj_list = []

            for value in sheet.iter_rows(
                    min_row=3, max_row=int(maximum_rows) + 1, min_col=2, max_col=5,
                    values_only=True):
                routes.append(value)
            bus = Bus.objects.get(bus_full_name=bus)

            for item in routes:
                a, b, c, d = item
                #     # print(f'Route(starting_place={a}, destination={b}, departure_time={c}, price={d})')
                obj_list.append(Route(bus=bus, starting_place=a, destination=b, time=c, price=d))
            #     # print(Route(a, b, c, d))
            #     result = Route.objects.create(bus=bus, starting_place=a, destination=b, time=c, price=d)
            Route.objects.bulk_create(obj_list)

        return HttpResponse('All Saved')
